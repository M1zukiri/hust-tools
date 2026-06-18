"""
XLSX/XLS/CSV 转 Markdown 转换器
使用 openpyxl 和 pandas 库
"""

from pathlib import Path
from typing import Optional, Dict, Any, List
import io

from converter_base import BaseConverter, ConversionResult, ConversionStatus


class XlsxConverter(BaseConverter):
    """Excel电子表格转Markdown转换器"""
    
    supported_extensions = ['xlsx', 'xls', 'csv', 'xlsm', 'xlsb']
    supported_mime_types = [
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'application/vnd.ms-excel',
        'text/csv'
    ]
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        super().__init__(config)
        self.max_rows = config.get('max_rows', 1000) if config else 1000
        self.max_cols = config.get('max_cols', 50) if config else 50
        self.include_sheet_names = config.get('include_sheet_names', True) if config else True
    
    def convert(self, file_path: Path, **options) -> ConversionResult:
        """
        将Excel/CSV文件转换为Markdown
        
        Args:
            file_path: 文件路径
            **options: 额外选项
                - max_rows: 最大行数限制
                - max_cols: 最大列数限制
                - include_sheet_names: 是否包含工作表名称
                - sheet_index: 指定工作表索引（仅Excel）
                
        Returns:
            ConversionResult: 转换结果
        """
        if not self.validate_file(file_path):
            return self._create_error_result(file_path, "文件验证失败", "xlsx")
        
        ext = file_path.suffix.lower()
        
        try:
            if ext == '.csv':
                return self._convert_csv(file_path, **options)
            else:
                return self._convert_excel(file_path, **options)
                
        except Exception as e:
            self.logger.error(f"转换失败: {e}")
            return self._create_error_result(file_path, str(e), "xlsx")
    
    def _convert_csv(self, file_path: Path, **options) -> ConversionResult:
        """
        转换CSV文件
        
        Args:
            file_path: CSV文件路径
            **options: 额外选项
            
        Returns:
            ConversionResult: 转换结果
        """
        try:
            import pandas as pd
        except ImportError:
            return self._create_error_result(
                file_path,
                "缺少依赖: pandas。请运行: pip install pandas",
                "csv"
            )
        
        try:
            # 检测编码
            encoding = options.get('encoding', 'utf-8')
            
            # 读取CSV
            df = pd.read_csv(
                str(file_path),
                encoding=encoding,
                nrows=options.get('max_rows', self.max_rows)
            )
            
            # 限制列数
            if len(df.columns) > options.get('max_cols', self.max_cols):
                df = df.iloc[:, :options.get('max_cols', self.max_cols)]
            
            md_content = self._dataframe_to_markdown(df)
            
            metadata = self.extract_metadata(file_path)
            metadata['rows'] = len(df)
            metadata['columns'] = len(df.columns)
            
            return ConversionResult(
                content=md_content,
                status=ConversionStatus.SUCCESS,
                source_format='csv',
                metadata=metadata,
                errors=[]
            )
            
        except Exception as e:
            return self._create_error_result(file_path, str(e), "csv")
    
    def _convert_excel(self, file_path: Path, **options) -> ConversionResult:
        """
        转换Excel文件
        
        Args:
            file_path: Excel文件路径
            **options: 额外选项
            
        Returns:
            ConversionResult: 转换结果
        """
        try:
            import openpyxl
        except ImportError:
            return self._create_error_result(
                file_path,
                "缺少依赖: openpyxl。请运行: pip install openpyxl",
                "xlsx"
            )
        
        try:
            wb = openpyxl.load_workbook(str(file_path), data_only=True)
            
            md_parts = []
            total_rows = 0
            total_cols = 0
            
            # 获取指定工作表或全部
            sheet_index = options.get('sheet_index', None)
            
            if sheet_index is not None:
                sheets = [wb.worksheets[sheet_index]] if sheet_index < len(wb.worksheets) else []
            else:
                sheets = wb.worksheets
            
            for idx, sheet in enumerate(sheets):
                if options.get('include_sheet_names', self.include_sheet_names):
                    md_parts.append(f"## 工作表: {sheet.title}\n")
                
                md_table = self._worksheet_to_markdown(
                    sheet,
                    max_rows=options.get('max_rows', self.max_rows),
                    max_cols=options.get('max_cols', self.max_cols)
                )
                
                if md_table.strip():
                    md_parts.append(md_table)
                    total_rows += sheet.max_row
                    total_cols = max(total_cols, sheet.max_column)
                
                if idx < len(sheets) - 1:
                    md_parts.append("\n---\n")
            
            metadata = self.extract_metadata(file_path)
            metadata['sheet_count'] = len(sheets)
            metadata['total_rows'] = total_rows
            metadata['max_columns'] = total_cols
            
            return ConversionResult(
                content='\n'.join(md_parts),
                status=ConversionStatus.SUCCESS,
                source_format='xlsx',
                metadata=metadata,
                errors=[]
            )
            
        except Exception as e:
            return self._create_error_result(file_path, str(e), "xlsx")
    
    def _worksheet_to_markdown(self, sheet, max_rows: int = 1000, max_cols: int = 50) -> str:
        """
        将工作表转换为Markdown表格
        
        Args:
            sheet: openpyxl工作表对象
            max_rows: 最大行数
            max_cols: 最大列数
            
        Returns:
            str: Markdown表格
        """
        rows = []
        row_count = min(sheet.max_row, max_rows)
        col_count = min(sheet.max_column, max_cols)
        
        for row_idx in range(1, row_count + 1):
            row_data = []
            for col_idx in range(1, col_count + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # 处理不同类型的值
                if value is None:
                    value = ""
                elif isinstance(value, (int, float)):
                    value = str(value)
                else:
                    value = str(value).strip()
                    # 转义管道符
                    value = value.replace('|', '\\|')
                    # 处理换行
                    value = value.replace('\n', ' ')
                
                row_data.append(value)
            
            rows.append(row_data)
        
        if not rows:
            return ""
        
        return self._build_markdown_table(rows)
    
    def _dataframe_to_markdown(self, df) -> str:
        """
        将pandas DataFrame转换为Markdown表格
        
        Args:
            df: DataFrame对象
            
        Returns:
            str: Markdown表格
        """
        rows = []
        
        # 表头
        header = [str(col) for col in df.columns]
        rows.append(header)
        
        # 数据行
        for _, row in df.iterrows():
            row_data = []
            for val in row:
                if pd.isna(val):
                    row_data.append("")
                else:
                    val_str = str(val).replace('|', '\\|').replace('\n', ' ')
                    row_data.append(val_str)
            rows.append(row_data)
        
        return self._build_markdown_table(rows)
    
    def _build_markdown_table(self, rows: List[List[str]]) -> str:
        """
        构建Markdown表格
        
        Args:
            rows: 行数据列表
            
        Returns:
            str: Markdown表格
        """
        if not rows:
            return ""
        
        md_lines = []
        
        # 表头
        md_lines.append('| ' + ' | '.join(rows[0]) + ' |')
        
        # 分隔符
        md_lines.append('|' + '|'.join(['---' for _ in rows[0]]) + '|')
        
        # 数据行
        for row in rows[1:]:
            # 确保列数一致
            while len(row) < len(rows[0]):
                row.append("")
            row = row[:len(rows[0])]
            
            md_lines.append('| ' + ' | '.join(row) + ' |')
        
        return '\n'.join(md_lines)


# 导入pandas用于类型检查
try:
    import pandas as pd
except ImportError:
    pd = None
